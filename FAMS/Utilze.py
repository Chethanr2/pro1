import openpyxl

def GetRowCount (File, sheetName):
    workbook = openpyxl.load_workbook(File)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def GetColCount(File, SheetName):
    workbook = openpyxl.load_workbook(File)
    sheet = workbook.get_sheet_by_name(SheetName)
    return(sheet.max_column)

