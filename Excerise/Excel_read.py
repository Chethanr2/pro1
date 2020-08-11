from selenium import webdriver
import openpyxl
import time

path=r"C:\Users\Chethan\Desktop\Names.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.active

rows = sheet.max_row
colms = sheet.max_column

print ( rows, colms)

for r in range(1,rows+1):
    for c in range(1,colms+1):
        print(sheet.cell(row=r,column=c).value,end="           ")

    print()



